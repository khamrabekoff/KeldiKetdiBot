"""Excel export with charts and formatting"""
import io
import database as db
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
import logging

logger = logging.getLogger(__name__)

# Colors
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TOTAL_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_monthly_report_excel(start_date, filename=None):
    """Create beautiful monthly report with charts"""
    try:
        rows = db.get_month_attendance_details(start_date)
        if not rows:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Oylik Hisobot"

        # Title
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = f"📊 Oylik Hisobot - {start_date.strftime('%B %Y')}"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Headers
        headers = ["Ism", "Sana", "Kelish", "Ketish", "Turi", "Tafsilot", "Jami ($)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        # Data
        row_num = 4
        employee_totals = {}

        for row in rows:
            rates = {
                'salary_type': row['salary_type'] if row['salary_type'] else 'tariff',
                'rate_n': row['rate_n'] if row['rate_n'] else 0,
                'rate_m': row['rate_m'] if row['rate_m'] else 0,
                'rate_k': row['rate_k'] if row['rate_k'] else 0,
                'rate_overtime': row['rate_overtime'] if row['rate_overtime'] else 0,
                'monthly_salary': row['monthly_salary'] if row['monthly_salary'] else 0,
                'overtime_hourly_rate': row['overtime_hourly_rate'] if row['overtime_hourly_rate'] else 0,
                'rate_per_minute': row['rate_per_minute'] if row['rate_per_minute'] else 0,
            }

            check_in = row['check_in']
            check_out = row['check_out']
            stype = rates['salary_type']
            stype_label = "Tarif" if stype == 'tariff' else ("Oylik" if stype == 'monthly' else "Minutlik")

            tafsilot = ""
            if check_in and check_out:
                from utils import calculate_wage
                _, _, breakdown = calculate_wage(check_in, check_out, rates)
                total = row['total_wage']
                if stype == 'per_minute':
                    total_mins = (check_out - check_in).total_seconds() / 60.0
                    tafsilot = f"{total_mins:.0f}min × ${rates['rate_per_minute']}"
                elif stype == 'monthly':
                    tafsilot = f"Base: ${breakdown.get('regular', 0):.2f}, OT: ${breakdown.get('ot', 0):.2f}"
                else:
                    tafsilot = f"N:{breakdown['n']:.0f} M:{breakdown['m']:.0f} K:{breakdown['k']:.0f} OT:{breakdown['ot']:.0f}"
            else:
                total = 0
                tafsilot = "Not finished"

            # Track totals by employee
            emp_name = row['full_name']
            if emp_name not in employee_totals:
                employee_totals[emp_name] = 0
            employee_totals[emp_name] += total

            # Row data
            ws.cell(row=row_num, column=1).value = emp_name
            ws.cell(row=row_num, column=2).value = row['date']
            ws.cell(row=row_num, column=3).value = check_in.strftime("%H:%M") if check_in else ""
            ws.cell(row=row_num, column=4).value = check_out.strftime("%H:%M") if check_out else ""
            ws.cell(row=row_num, column=5).value = stype_label
            ws.cell(row=row_num, column=6).value = tafsilot
            ws.cell(row=row_num, column=7).value = total

            # Format row
            for col in range(1, 8):
                cell = ws.cell(row=row_num, column=col)
                cell.border = BORDER
                if col == 7:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

            row_num += 1

        # Summary section
        row_num += 1
        ws.merge_cells(f'A{row_num}:G{row_num}')
        summary_title = ws[f'A{row_num}']
        summary_title.value = "UMUMIY TO'LOV"
        summary_title.font = TOTAL_FONT
        summary_title.fill = TOTAL_FILL
        summary_title.alignment = Alignment(horizontal="center")

        row_num += 1
        total_wage = sum(employee_totals.values())
        for emp_name, wage in sorted(employee_totals.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row_num, column=1).value = emp_name
            ws.cell(row=row_num, column=7).value = wage
            ws.cell(row=row_num, column=7).number_format = '$#,##0.00'
            ws.cell(row=row_num, column=7).font = TOTAL_FONT
            row_num += 1

        ws.cell(row=row_num, column=1).value = "JAMI:"
        ws.cell(row=row_num, column=7).value = total_wage
        ws.cell(row=row_num, column=7).number_format = '$#,##0.00'
        ws.cell(row=row_num, column=7).font = TOTAL_FONT
        ws.cell(row=row_num, column=7).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12

        # Create chart
        if len(employee_totals) > 0:
            chart_sheet = wb.create_sheet("Chart")

            # Prepare data for chart
            chart_sheet['A1'].value = "Ism"
            chart_sheet['B1'].value = "Jami ($)"

            row_num = 2
            for emp_name, wage in sorted(employee_totals.items(), key=lambda x: x[1], reverse=True):
                chart_sheet[f'A{row_num}'].value = emp_name
                chart_sheet[f'B{row_num}'].value = wage
                row_num += 1

            # Create chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Oylik to'lovlar"
            chart.y_axis.title = 'To\'lov ($)'
            chart.x_axis.title = 'Xodimlar'

            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=row_num-1)
            categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=row_num-1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 12
            chart.width = 20

            chart_sheet.add_chart(chart, "A5")

        # Save to bytes
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        return bio
    except Exception as e:
        logger.error(f"Error creating Excel report: {e}")
        return None


def create_employee_detailed_excel(user_id, days=30, filename=None):
    """Create detailed report for single employee"""
    try:
        from analytics import get_employee_stats
        from datetime import timedelta

        stats = get_employee_stats(user_id, days)
        if not stats:
            return None

        now = datetime.now()
        start_date = now.date() - timedelta(days=days)

        attendance = db.get_user_month_details(user_id, start_date)

        wb = Workbook()
        ws = wb.active
        ws.title = "Xodim Tafsiloti"

        # Header
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = f"📊 {stats['name']} - Tafsiliy Hisobot"
        title.font = HEADER_FONT
        title.fill = HEADER_FILL
        ws.row_dimensions[1].height = 20

        # Employee info
        row = 3
        ws[f'A{row}'].value = "Ism:"
        ws[f'B{row}'].value = stats['name']
        row += 1
        ws[f'A{row}'].value = "Telefon:"
        ws[f'B{row}'].value = stats['phone']
        row += 1
        ws[f'A{row}'].value = "Ish turi:"
        ws[f'B{row}'].value = stats['salary_type']
        row += 2

        # Stats summary
        ws[f'A{row}'].value = "STATISTIKA:"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        ws[f'A{row}'].value = "Ishlagan kunlar:"
        ws[f'B{row}'].value = stats['days_worked']
        row += 1
        ws[f'A{row}'].value = "Jami soatlar:"
        ws[f'B{row}'].value = stats['total_hours']
        ws[f'B{row}'].number_format = '0.0'
        row += 1
        ws[f'A{row}'].value = "O'rtacha soat/kun:"
        ws[f'B{row}'].value = stats.get('avg_hours_per_day', 0)
        ws[f'B{row}'].number_format = '0.0'
        row += 1
        ws[f'A{row}'].value = "Jami to'lov:"
        ws[f'B{row}'].value = stats['total_wage']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, size=11)
        row += 1
        ws[f'A{row}'].value = "O'rtacha to'lov/kun:"
        ws[f'B{row}'].value = stats['avg_wage_per_day']
        ws[f'B{row}'].number_format = '$#,##0.00'
        row += 2

        # Detail table
        row += 1
        headers = ["Sana", "Kelish", "Ketish", "Soatlar", "To'lov ($)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        row += 1
        for att in attendance:
            if att['check_in'] and att['check_out']:
                hours = (att['check_out'] - att['check_in']).total_seconds() / 3600

                ws[f'A{row}'].value = att['date']
                ws[f'B{row}'].value = att['check_in'].strftime("%H:%M")
                ws[f'C{row}'].value = att['check_out'].strftime("%H:%M")
                ws[f'D{row}'].value = hours
                ws[f'D{row}'].number_format = '0.00'
                ws[f'E{row}'].value = att['total_wage']
                ws[f'E{row}'].number_format = '$#,##0.00'

                row += 1

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        return bio
    except Exception as e:
        logger.error(f"Error creating employee report: {e}")
        return None
